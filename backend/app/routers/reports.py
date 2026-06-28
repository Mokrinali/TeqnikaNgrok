import math
import io
from datetime import datetime, date
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..dependencies import get_session, SessionContext
from .. import crud
from ..crud import calc_overtime_hours, calc_total_hours

router = APIRouter()

GEO_DAYS = {0: "ორშ", 1: "სამ", 2: "ოთხ", 3: "ხუთ", 4: "პარ", 5: "შაბ", 6: "კვი"}


def _fill(color: str):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def _font(bold=False, size=None, color="000000"):
    return Font(bold=bold, size=size, color=color)

def _align(h="general", wrap=False):
    return Alignment(horizontal=h, wrap_text=wrap)


def _build_buhaleria(ws, eq_logs, trip_logs, start: date):
    from calendar import month_name
    geo_months = ["","იანვარი","თებერვალი","მარტი","აპრილი","მაისი","ივნისი","ივლისი","აგვისტო","სექტემბერი","ოქტომბერი","ნოემბერი","დეკემბერი"]
    month_label = f"{geo_months[start.month]} {start.year}"

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"საერთო ჯამი — {month_label}"
    c.font = _font(bold=True, size=14, color="FFFFFF")
    c.fill = _fill("1A237E")
    c.alignment = _align("center")

    headers = ["კონტრაქტორი", "სახ.ნომ", "ტიპი", "დასახელება", "რაოდ.", "₾/ერთ.", "სულ ₾"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.fill = _fill("283593")
        cell.alignment = _align("center")
    ws.auto_filter.ref = "A2:G2"

    row = 3
    grand_total = 0.0

    eq_grouped: dict = {}
    for log in eq_logs:
        if not log.exit_timestamp:
            continue
        key = (log.contractor_id, log.contractor_name, log.equipment_id, log.equipment_name, float(log.daily_rate), log.plate_number or "")
        eq_grouped.setdefault(key, []).append(log)

    for key in sorted(eq_grouped, key=lambda k: (k[1], k[3])):
        contractor_name, eq_name, daily_rate, plate = key[1], key[3], key[4], key[5]
        logs = eq_grouped[key]
        work_days = len({l.entry_timestamp.date() for l in logs})
        total = work_days * daily_rate

        ws.cell(row=row, column=1, value=contractor_name)
        ws.cell(row=row, column=2, value=plate)
        c3 = ws.cell(row=row, column=3, value="ტექნიკა")
        c3.fill = _fill("E8F5E9")
        ws.cell(row=row, column=4, value=eq_name)
        ws.cell(row=row, column=5, value=work_days).alignment = _align("center")
        ws.cell(row=row, column=6, value=daily_rate).number_format = "#,##0"
        ws.cell(row=row, column=7, value=total).number_format = "#,##0"
        ws.cell(row=row, column=7).font = _font(bold=True)
        grand_total += total
        row += 1

    trip_grouped: dict = {}
    for t in trip_logs:
        key = (t.contractor_id, t.contractor_name, t.plate_number or "", t.work_type_name, float(t.price_per_trip))
        trip_grouped.setdefault(key, []).append(t)

    for key in sorted(trip_grouped, key=lambda k: (k[1], k[3])):
        contractor_name, plate, wt_name, price = key[1], key[2], key[3], key[4]
        trips = trip_grouped[key]
        total_trips = sum(t.trip_count for t in trips)
        total = sum(t.trip_count * float(t.price_per_trip) for t in trips)

        ws.cell(row=row, column=1, value=contractor_name)
        ws.cell(row=row, column=2, value=plate)
        c3 = ws.cell(row=row, column=3, value="რეისი")
        c3.fill = _fill("E3F2FD")
        ws.cell(row=row, column=4, value=wt_name)
        ws.cell(row=row, column=5, value=total_trips).alignment = _align("center")
        ws.cell(row=row, column=6, value=price).number_format = "#,##0"
        ws.cell(row=row, column=7, value=total).number_format = "#,##0"
        ws.cell(row=row, column=7).font = _font(bold=True)
        grand_total += total
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="სულ ჯამი").font = _font(bold=True, size=13)
    ws.cell(row=row, column=1).alignment = _align("right")
    c7 = ws.cell(row=row, column=7, value=grand_total)
    c7.font = _font(bold=True, size=13)
    c7.number_format = "#,##0"
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = _fill("C8E6C9")

    for w, col in zip([30, 12, 10, 28, 10, 12, 14], range(1, 8)):
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_contractor_sheet(ws, contractor_name, logs):
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = contractor_name
    c.font = _font(bold=True, size=14)
    c.alignment = _align("center")

    row = 3
    total_work_days = 0
    total_overtime = 0.0
    total_amount = 0.0

    eq_groups: dict = {}
    for log in logs:
        key = (log.equipment_id, log.equipment_name, log.plate_number, float(log.daily_rate))
        eq_groups.setdefault(key, []).append(log)

    for key in sorted(eq_groups, key=lambda k: k[1]):
        eq_name, plate, daily_rate = key[1], key[2], key[3]
        eq_logs = sorted(eq_groups[key], key=lambda l: l.entry_timestamp)

        label = eq_name
        if plate:
            label += f"  ·  {plate}"
        label += f"  ·  {int(daily_rate)} ₾/დღე"

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=label).font = _font(bold=True)
        ws.cell(row=row, column=1).fill = _fill("F0F0F2")
        row += 1

        col_headers = ["თარიღი", "კვ. დღე", "შემოსვლა", "გასვლა", "სულ სთ", "ზეგ. სთ"]
        for i, h in enumerate(col_headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = _font(bold=True)
            c.fill = _fill("D3D3D3")
            c.alignment = _align("center")
        row += 1

        for log in eq_logs:
            ws.cell(row=row, column=1, value=log.entry_timestamp.strftime("%d.%m.%Y"))
            ws.cell(row=row, column=2, value=GEO_DAYS[log.entry_timestamp.weekday()]).alignment = _align("center")
            ws.cell(row=row, column=3, value=log.entry_timestamp.strftime("%H:%M")).alignment = _align("center")
            ws.cell(row=row, column=4, value=log.exit_timestamp.strftime("%H:%M") if log.exit_timestamp else "ღია").alignment = _align("center")

            if log.exit_timestamp:
                hrs = round(calc_total_hours(log.entry_timestamp, log.exit_timestamp), 2)
                ws.cell(row=row, column=5, value=hrs).number_format = "0.00"
                ws.cell(row=row, column=5).alignment = _align("center")
                total_work_days += 1
                total_amount += daily_rate

            ot = calc_overtime_hours(log.exit_timestamp) if log.exit_timestamp else 0.0
            if ot > 0:
                c6 = ws.cell(row=row, column=6, value=ot)
                c6.number_format = "0.00"
                c6.fill = _fill("FFD580")
                c6.alignment = _align("center")
                total_overtime += ot
            row += 1

        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="სულ სამ. დღეები").font = _font(bold=True)
    ws.cell(row=row, column=5, value=total_work_days).font = _font(bold=True)
    ws.cell(row=row, column=5).alignment = _align("center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="სულ ზეგ. საათები").font = _font(bold=True)
    ws.cell(row=row, column=5, value=round(total_overtime, 2)).number_format = "0.00"
    ws.cell(row=row, column=5).font = _font(bold=True)
    ws.cell(row=row, column=5).alignment = _align("center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=f"ეკუთვნის:  {int(total_amount)} ₾")
    c.font = _font(bold=True, size=13)
    c.fill = _fill("C8E6C9")
    c.alignment = _align("center")

    for w, col in zip([15, 10, 12, 12, 12, 12], range(1, 7)):
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_tabeli(ws, logs, start: date, end: date):
    days = (end - start).days + 1
    ws.cell(row=1, column=1, value="კონტრაქტორი / ტექნიკა").font = _font(bold=True)

    for d in range(days):
        dt = datetime(start.year, start.month, start.day + d) if False else \
             date(start.year, start.month, 1)
        from datetime import timedelta
        dt = start + timedelta(days=d)
        c = ws.cell(row=1, column=d + 2, value=dt.day)
        c.font = _font(bold=True)
        c.alignment = _align("center")
        if dt.weekday() >= 5:
            c.fill = _fill("D3D3D3")

    sum_col = days + 2
    for i, h in enumerate(["სულ დღეები", "სულ საათები", "სულ ზეგ., სთ"]):
        c = ws.cell(row=1, column=sum_col + i, value=h)
        c.font = _font(bold=True)
        c.fill = _fill("C6EFCE")
        c.alignment = _align("center", wrap=True)

    groups: dict = {}
    for log in logs:
        key = (log.contractor_name, log.equipment_id, log.equipment_name)
        groups.setdefault(key, []).append(log)

    row = 2
    from datetime import timedelta
    for key in sorted(groups, key=lambda k: (k[0], k[2])):
        contractor_name, eq_name = key[0], key[2]
        ws.cell(row=row, column=1, value=f"{contractor_name} — {eq_name}")

        work_days, total_hours, total_overtime = 0, 0.0, 0.0
        for d in range(days):
            dt = start + timedelta(days=d)
            log = next((l for l in groups[key] if l.entry_timestamp.date() == dt and l.exit_timestamp), None)
            if not log:
                continue
            ot = calc_overtime_hours(log.exit_timestamp)
            c = ws.cell(row=row, column=d + 2, value=1)
            c.fill = _fill("FFD580") if ot > 0 else _fill("90EE90")
            c.alignment = _align("center")
            work_days += 1
            total_hours += calc_total_hours(log.entry_timestamp, log.exit_timestamp)
            total_overtime += ot

        ws.cell(row=row, column=sum_col, value=work_days).alignment = _align("center")
        ws.cell(row=row, column=sum_col + 1, value=round(total_hours, 2)).number_format = "0.00"
        ws.cell(row=row, column=sum_col + 2, value=total_overtime).number_format = "0.00"
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions[get_column_letter(sum_col)].width = 13
    ws.column_dimensions[get_column_letter(sum_col + 1)].width = 13
    ws.column_dimensions[get_column_letter(sum_col + 2)].width = 14


def _build_shedegebi(ws, logs, start: date):
    geo_months = ["","იანვარი","თებერვალი","მარტი","აპრილი","მაისი","ივნისი","ივლისი","აგვისტო","სექტემბერი","ოქტომბერი","ნოემბერი","დეკემბერი"]
    month_label = f"{geo_months[start.month]} {start.year}"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"შედეგები — {month_label}"
    c.font = _font(bold=True, size=14, color="FFFFFF")
    c.fill = _fill("1B5E20")
    c.alignment = _align("center")

    headers = ["კონტრ. / ტექნიკა", "სულ დღეები", "სულ საათები", "სულ ზეგ., სთ", "ეკუთვნის ₾"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill("2E7D32")
        c.alignment = _align("center")

    groups: dict = {}
    for log in logs:
        key = (log.contractor_name, log.equipment_id, log.equipment_name)
        groups.setdefault(key, []).append(log)

    data_row = 3
    g_days = g_hours = g_ot = g_amount = 0.0

    for i, key in enumerate(sorted(groups, key=lambda k: (k[0], k[2]))):
        eq_logs = [l for l in groups[key] if l.exit_timestamp]
        work_days = len({l.entry_timestamp.date() for l in eq_logs})
        hours = sum(calc_total_hours(l.entry_timestamp, l.exit_timestamp) for l in eq_logs)
        overtime = sum(calc_overtime_hours(l.exit_timestamp) for l in eq_logs)
        amount = sum(float(l.daily_rate) for l in eq_logs)

        g_days += work_days; g_hours += hours; g_ot += overtime; g_amount += amount

        bg = "F1F8E9" if i % 2 == 0 else "FFFFFF"
        ws.cell(row=data_row, column=1, value=f"{key[0]} — {key[2]}")
        ws.cell(row=data_row, column=2, value=work_days).alignment = _align("center")
        ws.cell(row=data_row, column=3, value=round(hours, 2)).number_format = "0.00"
        ws.cell(row=data_row, column=4, value=overtime).number_format = "0.00"
        ws.cell(row=data_row, column=5, value=amount).number_format = "#,##0"
        for col in range(1, 6):
            ws.cell(row=data_row, column=col).fill = _fill(bg)
        data_row += 1

    ws.cell(row=data_row, column=1, value="სულ ჯამი").font = _font(bold=True)
    ws.cell(row=data_row, column=2, value=int(g_days)).font = _font(bold=True)
    ws.cell(row=data_row, column=3, value=round(g_hours, 2)).number_format = "0.00"
    ws.cell(row=data_row, column=4, value=g_ot).number_format = "0.00"
    ws.cell(row=data_row, column=5, value=g_amount).number_format = "#,##0"
    for col in range(1, 6):
        ws.cell(row=data_row, column=col).fill = _fill("C8E6C9")
        ws.cell(row=data_row, column=col).font = _font(bold=True)

    for w, col in zip([42, 14, 15, 15, 16], range(1, 6)):
        ws.column_dimensions[get_column_letter(col)].width = w


@router.get("/export-excel")
async def export_excel(
    from_dt: date | None = None,
    to_dt: date | None = None,
    contractor_id: int | None = None,
    site_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    session: SessionContext = Depends(get_session),
):
    import calendar
    today = date.today()
    start = from_dt or date(today.year, today.month, 1)
    if to_dt:
        end = to_dt
    else:
        last_day = calendar.monthrange(start.year, start.month)[1]
        end = date(start.year, start.month, last_day)

    effective_site = site_id if session.is_admin else session.current_site_id

    start_dt = datetime(start.year, start.month, start.day, 0, 0, 0)
    end_dt = datetime(end.year, end.month, end.day, 23, 59, 59)

    eq_logs = await crud.get_logs(db, start_dt, end_dt, effective_site, contractor_id)
    trip_logs = await crud.get_trip_logs(db, start, end, effective_site, contractor_id)

    wb = Workbook()
    wb.remove(wb.active)

    ws_buh = wb.create_sheet("ბუღალტერია")
    _build_buhaleria(ws_buh, eq_logs, trip_logs, start)

    grouped: dict = {}
    for log in eq_logs:
        key = (log.contractor_id, log.contractor_name)
        grouped.setdefault(key, []).append(log)

    for key in sorted(grouped, key=lambda k: k[1]):
        sheet_name = key[1][:25]
        ws_c = wb.create_sheet(sheet_name)
        _build_contractor_sheet(ws_c, key[1], grouped[key])

    ws_tab = wb.create_sheet("ტაბელი")
    _build_tabeli(ws_tab, eq_logs, start, end)

    ws_shed = wb.create_sheet("შედეგები")
    _build_shedegebi(ws_shed, eq_logs, start)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"tabeli_{start.year}_{start.month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-trips-excel")
async def export_trips_excel(
    from_dt: date | None = None,
    to_dt: date | None = None,
    contractor_id: int | None = None,
    site_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    session: SessionContext = Depends(get_session),
):
    today = date.today()
    start = from_dt or today
    end = to_dt or today
    effective_site = site_id if session.is_admin else session.current_site_id
    trips = await crud.get_trip_logs(db, start, end, effective_site, contractor_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "რეისები"

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"რეისები: {start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}"
    c.font = _font(bold=True, size=13)

    headers = ["თარიღი", "კონტრაქტორი", "ნომერი", "სამუშაო", "რეისები", "ფასი/რეისი", "ჯამი"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = _font(bold=True)
        c.fill = _fill("D3D3D3")

    row = 3
    for t in sorted(trips, key=lambda x: x.date):
        total = float(t.trip_count) * float(t.price_per_trip)
        ws.cell(row=row, column=1, value=t.date.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=2, value=t.contractor_name)
        ws.cell(row=row, column=3, value=t.plate_number or "")
        ws.cell(row=row, column=4, value=t.work_type_name)
        ws.cell(row=row, column=5, value=t.trip_count)
        ws.cell(row=row, column=6, value=float(t.price_per_trip))
        ws.cell(row=row, column=7, value=total).font = _font(bold=True)
        row += 1

    grand = sum(float(t.trip_count) * float(t.price_per_trip) for t in trips)
    ws.cell(row=row, column=6, value="სულ:").font = _font(bold=True)
    c = ws.cell(row=row, column=7, value=grand)
    c.font = _font(bold=True)
    c.fill = _fill("90EE90")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"reisebi_{start.strftime('%Y_%m_%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
